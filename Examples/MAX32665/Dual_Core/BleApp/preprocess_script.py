import pandas as pd
from openpyxl import load_workbook
import openpyxl
import time
from openpyxl.styles import PatternFill
class ProcessedData:
    def __init__(self, timestamp:float, x_val:int, y_val:int, z_val:int):
        self.timestamp = timestamp
        self.x_val = x_val
        self.y_val = y_val
        self.z_val = z_val

MAX_MOTION_PACK_LEN = 50
# Main function to process Excel file
def process_excel(input_file, output_file):
    new_file_row = 1
    new_file_column = 1
    current_motion_data:list[ProcessedData] = []
    pre_padding_data:list[ProcessedData] = []
    motion_started = False
    motion_ended = False

    raw_data_workbook = load_workbook(input_file)
    raw_data_sheet = raw_data_workbook.active
    max_row = raw_data_sheet.max_row
    max_column = raw_data_sheet.max_column

    try:
        new_file_workbook = load_workbook(output_file)
        new_file_sheet = new_file_workbook.active
    except:
        new_file_workbook  = openpyxl.Workbook()
        new_file_workbook.save(output_file)
        new_file_sheet = new_file_workbook.active
    
    average_motion_duration = 0
    row_count = 0
    for row in raw_data_sheet.iter_rows(min_row=1, max_row= max_row, min_col=1, max_col=max_column, values_only= True ):
        i = 0
        while i < len(row):
            if 0 == row[i]:
                if 0 != row[i+1] and None != row[i+2] and None != row[i+3] and None!= row[i+4]:
                    current_motion_data.append(ProcessedData(row[i+1],row[i+2],row[i+3],row[i+4]))
                motion_started = True
            else:
                if motion_started == True:
                    motion_ended = True
                    break
                if 0 != row[i+1] and None != row[i+2] and None != row[i+3] and None!= row[i+4]:
                    if len(pre_padding_data) >=30:
                        pre_padding_data.pop(0)
                    pre_padding_data.append(ProcessedData(row[i+1],row[i+2],row[i+3],row[i+4]))
            i = i+5

        if True == motion_started and True == motion_ended:
            motion_started = False
            motion_ended = False
            # Sampling rate is 10ms for each packet.
            # 15*10 = 250, 50*10 = 500ms.
            # Motion duration must be between this time interval [250ms,500ms].Less than this will not be taken as valid data
            print(len(current_motion_data))

            if len(current_motion_data) >= 20:
                row_count = row_count + 1
                average_motion_duration = average_motion_duration + len(current_motion_data)
                new_file_column = 1
                merged_data:list[ProcessedData] = []
                required_padding_pack_len = MAX_MOTION_PACK_LEN-len(current_motion_data)
                idx = 0

                # Alternative Approach 1 : Zero prepadding + Real data prepadding 
                if required_padding_pack_len > 0:
                    pre_padding_len = len(pre_padding_data)
                    if pre_padding_len >= required_padding_pack_len:
                        for pack in pre_padding_data[pre_padding_len-required_padding_pack_len:]:
                            merged_data.append(pack)
                    else:
                        zero_padding_size = required_padding_pack_len - pre_padding_len
                        while idx < zero_padding_size:
                            merged_data.append(ProcessedData(0,0,0,0))
                            idx = idx+1
                        for pack in pre_padding_data:
                            merged_data.append(pack)

                # Alternative Approach 2 : Full zero padding
                # if required_padding_pack_len > 0:        
                #     while idx < required_padding_pack_len:
                #         merged_data.append(ProcessedData(0,0,0,0))
                #         idx = idx+1
                motion_data_start_index = len(merged_data)
                for pack in current_motion_data:
                    merged_data.append(pack)
                    if len(merged_data) == 50:
                        break

                back_fill_for_motion_data = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green color
                for merged_pack in merged_data:
                    if motion_data_start_index > 0:
                        motion_data_start_index = motion_data_start_index - 1
                        new_file_sheet.cell(row = new_file_row, column= new_file_column * 4 - 3, value=merged_pack.timestamp)
                        new_file_sheet.cell(row = new_file_row, column= new_file_column * 4 - 2, value=merged_pack.x_val)
                        new_file_sheet.cell(row = new_file_row, column= new_file_column * 4 - 1, value=merged_pack.y_val)
                        new_file_sheet.cell(row = new_file_row, column= new_file_column * 4, value=merged_pack.z_val)
                    else:
                        new_file_sheet.cell(row = new_file_row, column= new_file_column * 4 - 3, value=merged_pack.timestamp).fill = back_fill_for_motion_data
                        new_file_sheet.cell(row = new_file_row, column= new_file_column * 4 - 2, value=merged_pack.x_val).fill = back_fill_for_motion_data
                        new_file_sheet.cell(row = new_file_row, column= new_file_column * 4 - 1, value=merged_pack.y_val).fill = back_fill_for_motion_data
                        new_file_sheet.cell(row = new_file_row, column= new_file_column * 4, value=merged_pack.z_val).fill = back_fill_for_motion_data

                    new_file_column = new_file_column + 1
                    if new_file_column > MAX_MOTION_PACK_LEN+1:
                        break

                if len(current_motion_data) != 0:
                    new_file_row = new_file_row +1       
            pre_padding_data.clear()
            current_motion_data.clear()

    print(average_motion_duration/row_count)
    new_file_workbook.save(output_file)         

if __name__ == "__main__":
    input_file_name = input("Input filename:")
    output_file_name =  "processedTest"+input_file_name
    process_excel(input_file_name, output_file_name)