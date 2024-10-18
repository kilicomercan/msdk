import openpyxl
from datetime import datetime
import threading
class BleFile:
    def __init__(self):
        self.mark_sign = 'None'
        self.mark_enable = False
        self.filename = ""
        self.workbook = None
        self.sheet = None
        self.current_row = 1
        self.__current_packet = 1
        self.is_recording = False
        self.file_lock = threading.Lock()
        self.__data_collection_path = ""

    def generate_filename(self):
        # Generate a filename with the current time of the day
        current_time = datetime.now().strftime("%H-%M-%S")
        filename = current_time+"_DataSet"
        return f"{filename}.xlsx"

    def open_excel_file(self):
        try:
            # Open the Excel file with the generated filename
            workbook = openpyxl.load_workbook(self.__data_collection_path+self.filename)
            print(f"Opened file: {self.filename}")
            return workbook
        except FileNotFoundError:
            print(f"File {self.filename} not found. Creating a new one.")
            workbook = openpyxl.Workbook()
            workbook.save(self.__data_collection_path+self.filename)
            return workbook
    
    def write_packet(self,flag:int, timestamp:float, x_val:int, y_val:int, z_val:int):
        # Write a packet to the current row
        self.sheet.cell(row=self.current_row, column=self.__current_packet * 5 - 4, value = flag)
        self.sheet.cell(row=self.current_row, column=self.__current_packet * 5 - 3, value=timestamp)
        self.sheet.cell(row=self.current_row, column=self.__current_packet * 5 - 2, value=x_val)
        self.sheet.cell(row=self.current_row, column=self.__current_packet * 5 - 1, value=y_val)
        self.sheet.cell(row=self.current_row, column=self.__current_packet * 5, value=z_val)
        
        # If it's the last packet in the row, write the mark sign
        if 0 == self.__current_packet % 100:
            if self.mark_enable:
                self.sheet.cell(row=self.current_row, column=501, value=self.mark_sign)
            self.current_row += 1
            self.__current_packet = 1
        else:
            self.__current_packet += 1

        # Save the workbook after writing each packet
        # self.workbook.save(self.filename)
    
    def create_new_file(self):
        self.filename = self.generate_filename()
        self.workbook = self.open_excel_file()
        self.sheet = self.workbook.active
        self.current_row = 1
        self.__current_packet = 1

    def conf_mark_sign(self, mark_enable = False, mark_sign = None ):
        self.mark_enable = mark_enable
        self.mark_sign = mark_sign

    def close_file(self):
        if self.workbook != None:
            self.workbook.save(self.filename)
        self.mark_sign = 'None'
        self.mark_enable = False
        self.filename = ""
        self.workbook = None
        self.sheet = None
        self.current_row = 1
        self.__current_packet = 1

    def __self_test(self):
        for i in range(200):  # Example to write 200 packets
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            x_val = i
            y_val = i * 2
            z_val = i * 3
            self.write_packet(timestamp, x_val, y_val, z_val)
