import openpyxl
from openpyxl.styles import PatternFill

# Load the pre file. We will read the data from this file and 
# save it to a new file after applying sliding window method
# to pre padding zeros.
file_path = input("Provide file name?")
pre_file = openpyxl.load_workbook(file_path)
pre_file_sheet = pre_file.active  # Assumes you are working with the first sheet; you can also specify the sheet name


post_workbook = openpyxl.Workbook()
post_workbook_sheet = post_workbook.active

# Define the RGB color for green (assuming 'green' means solid fill of #00FF00)
# green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
# green_rgb = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid") 
green_rgb = 'FF00FF00'
# Iterate over all rows and first 200 columns
for row in pre_file_sheet.iter_rows(min_row=1, max_col=200):
    pre_padding = []
    motion_data = []
    for cell in row:
        fill = cell.fill
        if fill is not None and fill.fgColor is not None and fill.fgColor.rgb != green_rgb:
            pre_padding.append(cell.value)
        else:
            motion_data.append(cell.value)


                



post_workbook.save("Slided" + file_path)