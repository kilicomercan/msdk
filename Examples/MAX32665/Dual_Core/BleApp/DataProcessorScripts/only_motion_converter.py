import openpyxl
from openpyxl.styles import PatternFill

# Load the Excel file
file_path = input("Provide file name?")
wb = openpyxl.load_workbook(file_path)
ws = wb.active  # Assumes you are working with the first sheet; you can also specify the sheet name

# Define the RGB color for green (assuming 'green' means solid fill of #00FF00)
# green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
# green_rgb = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid") 
green_rgb = 'FF00FF00'
# Iterate over all rows and first 200 columns
for row in ws.iter_rows(min_row=1, max_col=200):
    for cell in row:
        fill = cell.fill
        if fill is not None and fill.fgColor is not None and fill.fgColor.rgb != green_rgb:
            cell.value = 0  # Change the cell value as needed
       
# Save the modified workbook
output_file = "modified_" + file_path
wb.save(output_file)

print(f"Modified file saved as {output_file}")
